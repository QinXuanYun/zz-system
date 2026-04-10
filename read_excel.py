# -*- coding: utf-8 -*-
import openpyxl
import json
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(
    r'D:\workspace\huiyi_pro\xingjian\指标、阈值及数据0408.xlsx',
    data_only=True
)

print("Sheets:", wb.sheetnames)

# Build indicator metadata from Sheet1
ws_thresh = wb.worksheets[0]
indicators_meta = []

print("\n=== Sheet1 (Thresholds) ===")
for r in range(2, ws_thresh.max_row + 1):
    row = [ws_thresh.cell(r, c).value for c in range(1, ws_thresh.max_column + 1)]
    if row[0] is None:
        break
    seq = int(row[0])
    name = str(row[1]) if row[1] else ''
    method = str(row[2]) if row[2] else ''
    threshold_str = str(row[3]) if row[3] else ''
    print(f"  X{seq}: {name}")
    print(f"       method: {method}")
    print(f"       threshold: {threshold_str}")
    indicators_meta.append({
        "id": f"X{seq}",
        "name": name,
        "method": method,
        "threshold_str": threshold_str
    })

# Map Excel sheet index to major id
major_map = [
    ("feixingqi", "航空装备制造类", 1),
    ("jidian", "机电一体化类", 2),
    ("qiche", "汽车检测与维修技术", 3),
    ("ruanjian", "计算机网络技术", 4),
    ("jisuanji", "软件技术", 5),
]

years = ["2020学年", "2021学年", "2022学年"]

all_data = {}

for mid, mname, sheet_idx in major_map:
    ws = wb.worksheets[sheet_idx]
    print(f"\n=== {mname} ===")
    for year_idx, year in enumerate(years):
        mdata_raw = {}
        mdata_score = {}
        print(f"\n{year}:")
        for r in range(2, 17):  # rows 2-16
            row = [ws.cell(r, c).value for c in range(1, 6)]
            if row[0] is None:
                break
            seq = int(row[0])
            raw_val = row[3]
            score_val = row[4]
            key = f"X{seq}"
            mdata_raw[key] = raw_val
            mdata_score[key] = score_val
            # Show first few
            if seq <= 10:
                ind_name = indicators_meta[seq-1]["name"] if seq <= len(indicators_meta) else f"X{seq}"
                print(f"  X{seq}({ind_name}): raw={raw_val}, score={score_val}")
        
        if year not in all_data:
            all_data[year] = {}
        
        total = sum(mdata_score.get(f"X{i}", 0) for i in range(1, 11))
        all_data[year][mid] = {
            "raw": mdata_raw,
            "score": mdata_score,
            "total": total,
            "scoreRate": round(total / 50.0, 4)
        }
