# -*- coding: utf-8 -*-
import openpyxl
import json
import os
import sys

# Ensure stdout uses UTF-8
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(
    r'D:\workspace\huiyi_pro\xingjian\指标、阈值及数据0408.xlsx',
    data_only=True
)

print("Sheet names:", wb.sheetnames)
print()

# Major names from sheet order
major_sheet_names = [
    '航空装备制造类',
    '机电一体化类',
    '汽车检测与维修技术',
    '计算机网络技术',
    '软件技术'
]

# Build indicator metadata from Sheet1 (thresholds)
ws_thresh = wb.worksheets[0]
indicators_meta = {}
indicator_names = []

print("=== Sheet1 (Thresholds) ===")
for r in range(2, ws_thresh.max_row + 1):
    row = [ws_thresh.cell(r, c).value for c in range(1, ws_thresh.max_column + 1)]
    if row[0] is None:
        break
    seq = int(row[0])
    name = str(row[1]) if row[1] else ''
    method = str(row[2]) if row[2] else ''
    threshold_str = str(row[3]) if row[3] else ''
    print(f"  X{seq}: {name} | method: {method} | threshold: {threshold_str}")
    indicator_names.append(name)

print()
print("Indicator names (in order):", indicator_names)
print()

# Collect all data
years = ['2020学年', '2021学年', '2022学年']
majors_data = {}
major_ids = ['feixingqi', 'jidian', 'qiche', 'ruanjian', 'jisuanji']
major_real_names = major_sheet_names

for year_idx, year in enumerate(years):
    majors_data[year] = {}
    for midx, mid in enumerate(major_ids):
        ws = wb.worksheets[midx + 1]  # skip threshold sheet
        mdata = {}
        for r in range(2, 17):  # rows 2-16 are indicators
            row = [ws.cell(r, c).value for c in range(1, 6)]
            if row[0] is None:
                break
            seq = int(row[0])
            raw_val = row[3]  # column D = raw value
            score_val = row[4]  # column E = score
            mdata[f'X{seq}'] = {
                'raw': raw_val,
                'score': score_val
            }
        majors_data[year][mid] = mdata

print("=== Data preview ===")
for year, majors in majors_data.items():
    print(f"\n{year}:")
    for mid, mdata in majors.items():
        x2_raw = mdata.get('X2', {}).get('raw', 'N/A')
        x2_score = mdata.get('X2', {}).get('score', 'N/A')
        x1_raw = mdata.get('X1', {}).get('raw', 'N/A')
        print(f"  {mid}: X1={x1_raw}, X2={x2_raw} (score={x2_score})")

# Show ALL X2 values
print("\n=== X2 (生师比) values ===")
for year, majors in majors_data.items():
    for mid, mdata in majors.items():
        x2_raw = mdata.get('X2', {}).get('raw', 'N/A')
        print(f"  {year} {mid}: {x2_raw}")

# Build the new indicators.json
new_meta = {
    "meta": {
        "school": "信息与机电工程系",
        "years": years,
        "indicators": [
            {"id": "X1", "name": "招生计划完成率", "weight": 5, "method": "(实际录取数/招生计划数)*100%", "thresholds": {"red": (0, 0.85), "yellow": (0.85, 0.90), "green": (0.90, 999)}, "higher_is_better": True, "format": "pct"},
            {"id": "X2", "name": "生师比", "weight": 3, "method": "折合在校生数/折合专任教师数", "thresholds": {"green": 18, "yellow": 22, "red": 999}, "higher_is_better": False, "format": "ratio"},
            {"id": "X3", "name": "课程优良率", "weight": 3, "method": "学生评教分数/专业课程门数", "thresholds": {"red": (0, 0.70), "yellow": (0.70, 0.85), "green": (0.85, 999)}, "higher_is_better": True, "format": "pct"},
            {"id": "X4", "name": "技能证书通过率", "weight": 4, "method": "获得职业资格证书学生数/学年生均学生数", "thresholds": {"red": (0, 0.60), "yellow": (0.60, 0.75), "green": (0.75, 999)}, "higher_is_better": True, "format": "pct"},
            {"id": "X5", "name": "毕业率", "weight": 3, "method": "毕业学生数/招生总数", "thresholds": {"red": (0, 0.92), "yellow": (0.92, 0.95), "green": (0.95, 999)}, "higher_is_better": True, "format": "pct"},
            {"id": "X6", "name": "就业去向落实率", "weight": 5, "method": "落实去向学生数/毕业生总数", "thresholds": {"red": (0, 0.92), "yellow": (0.92, 0.96), "green": (0.96, 999)}, "higher_is_better": True, "format": "pct"},
            {"id": "X7", "name": "专业相关度", "weight": 4, "method": "专业对口岗位学生数/总岗位学生数", "thresholds": {"red": (0, 0.68), "yellow": (0.68, 0.70), "green": (0.70, 999)}, "higher_is_better": True, "format": "pct"},
            {"id": "X8", "name": "在校生满意度", "weight": 4, "method": "非常满意人数/总问卷人数", "thresholds": {"red": (0, 0.91), "yellow": (0.91, 0.95), "green": (0.95, 999)}, "higher_is_better": True, "format": "pct"},
            {"id": "X9", "name": "毕业生满意度", "weight": 4, "method": "满意人数/总问卷人数", "thresholds": {"red": (0, 0.92), "yellow": (0.92, 0.95), "green": (0.95, 999)}, "higher_is_better": True, "format": "pct"},
            {"id": "X10", "name": "企业订单学生占比", "weight": 4, "method": "企业订单培养学生数/年度招生总数", "thresholds": {"red": (0, 0.08), "yellow": (0.08, 0.15), "green": (0.15, 999)}, "higher_is_better": True, "format": "pct"},
        ],
        "majors": [
            {"id": "feixingqi", "name": "航空装备制造类", "fullName": "航空装备制造类"},
            {"id": "jidian", "name": "机电一体化技术", "fullName": "机电一体化技术"},
            {"id": "qiche", "name": "汽车检测与维修技术", "fullName": "汽车检测与维修技术"},
            {"id": "ruanjian", "name": "软件技术", "fullName": "软件技术"},
            {"id": "jisuanji", "name": "计算机网络技术", "fullName": "计算机网络技术(含中高职贯通)"},
        ]
    },
    "data": {}
}

# Map Excel columns to our indicator IDs
# X1-X10 from Excel rows 2-11 (对应 indicators X1-X10)
# We need to figure out which Excel row corresponds to which indicator

# From the preview:
# Row 2: X1=招生计划完成率, raw=0.6889, score=3.44 (航空装备)
# Row 3: X2=生师比, raw=18, score=3 (航空装备)
# Row 4: X3=课程优良率, raw=0.88, score=2.64 (航空装备)
# Row 5: X4=技能证书通过率, raw=0.5143, score=1.54 (航空装备)
# Row 6: X5=就业率, raw=0.96, score=2.88 (航空装备)
# Row 7: X6=就业去向落实率, raw=1.0, score=5 (航空装备)


# Row 8: X7=专业相关度, raw=0.71, score=4 (航空装备)
# Row 9: X8=在校生满意度, raw=0.945, score=3.96 (航空装备)
# Row 10: X9=毕业生满意度, raw=0.8703, score=3.54 (航空装备)
# Row 11: X10=企业订单学生占比, raw=0.3804, score=4 (航空装备)

# Rows 12-15 are师资 and 科研 metrics (双师型, 高级职称, 高技术技能, 师均论文, 教师企业实践)

# This means: Excel rows 2-11 = X1-X10 which is exactly our 10 indicators!
# The mapping is 1:1

print("\n=== Building new data structure ===")
# Build new data structure
for year_idx, year in enumerate(years):
    new_meta["data"][year] = {}
    for midx, mid in enumerate(major_ids):
        ws = wb.worksheets[midx + 1]
        mdata_raw = {}
        mdata_score = {}
        for r in range(2, 17):  # rows 2-16
            row = [ws.cell(r, c).value for c in range(1, 6)]
            if row[0] is None:
                break
            seq = int(row[0])
            raw_val = row[3]
            score_val = row[4]
            key = f'X{seq}'
            mdata_raw[key] = raw_val
            # Score from Excel
            mdata_score[key] = score_val
        
        # Calculate total from Excel scores
        total = sum(mdata_score.get(f'X{i}', 0) for i in range(1, 11))
        score_rate = total / 50.0  # 10 indicators * 5 max score each
        
        new_meta["data"][year][mid] = {
            "raw": mdata_raw,
            "score": mdata_score,
            "total": total,
            "scoreRate": score_rate
        }

# Print X2 values to verify
print("\nX2 (生师比) values in new structure:")
for year, majors in new_meta["data"].items():
    for mid, mdata in majors.items():
        x2 = mdata["raw"].get("X2", "N/A")
        print(f"  {year} {mid}: {x2}")

# Save to file
output_path = r'C:\Users\cuizh\.openclaw\workspace-piter\website_zyzd\data\indicators.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(new_meta, f, ensure_ascii=False, indent=2)

print(f"\nSaved to {output_path}")
