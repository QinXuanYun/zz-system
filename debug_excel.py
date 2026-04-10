# -*- coding: utf-8 -*-
import openpyxl
import json
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'D:\workspace\huiyi_pro\xingjian\指标、阈值及数据0408.xlsx', data_only=True)

print('Sheets:', wb.sheetnames)

ws1 = wb.worksheets[0]
print('\n=== Sheet1 (Thresholds) ===')
for r in range(1, ws1.max_row+1):
    print(f'Row {r}:', [ws1.cell(r,c).value for c in range(1, ws1.max_column+1)])

print('\n=== Major Sheet 1 (航空装备制造类) ===')
ws2 = wb.worksheets[1]
for r in range(1, ws2.max_row+1):
    print(f'Row {r}:', [ws2.cell(r,c).value for c in range(1, ws2.max_column+1)])

print('\n=== Major Sheet 2 (机电一体化类) ===')
ws3 = wb.worksheets[2]
for r in range(1, ws3.max_row+1):
    print(f'Row {r}:', [ws3.cell(r,c).value for c in range(1, ws3.max_column+1)])

print('\n=== Major Sheet 3 (汽车检测与维修技术) ===')
ws4 = wb.worksheets[3]
for r in range(1, ws4.max_row+1):
    print(f'Row {r}:', [ws4.cell(r,c).value for c in range(1, ws4.max_column+1)])

print('\n=== Major Sheet 4 (计算机网络技术) ===')
ws5 = wb.worksheets[4]
for r in range(1, ws5.max_row+1):
    print(f'Row {r}:', [ws5.cell(r,c).value for c in range(1, ws5.max_column+1)])

print('\n=== Major Sheet 5 (软件技术) ===')
ws6 = wb.worksheets[5]
for r in range(1, ws6.max_row+1):
    print(f'Row {r}:', [ws6.cell(r,c).value for c in range(1, ws6.max_column+1)])
