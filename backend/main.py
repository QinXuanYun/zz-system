"""
专业发展智诊系统 - FastAPI Backend (v3)
With Database support, JWT authentication
"""
import json
import os
import math
import re
import shutil
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Any

from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

import jwt
import openpyxl
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# Import database
from database import (
    get_db, get_years, get_year_data, import_excel_data,
    IndicatorMeta, SessionLocal, init_db
)

# Register Chinese font for PDF - Support Windows, Linux, macOS
def register_chinese_font():
    """Try to register a Chinese font for PDF generation."""
    # Get the project root directory
    project_root = Path(__file__).parent.parent
    fonts_dir = project_root / "fonts"
    
    font_paths = []
    
    # Add fonts from project fonts directory
    if fonts_dir.exists():
        for f in fonts_dir.iterdir():
            if f.suffix.lower() in ['.ttf', '.otf', '.ttc']:
                font_paths.append(str(f))
    
    # Add system font paths
    font_paths.extend([
        r"C:\Windows\Fonts\msyh.ttc",
        r"C:\Windows\Fonts\simhei.ttf",
        r"C:\Windows\Fonts\simsun.ttc",
        r"C:\Windows\Fonts\msyhbd.ttc",
        "/usr/share/fonts/wqy-microhei/wqy-microhei.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/System/Library/Fonts/PingFang.ttc",
        "/Library/Fonts/Arial Unicode.ttf",
    ])
    
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                font_name = os.path.basename(font_path).split('.')[0]
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                print(f"[FONT] Successfully registered: {font_name} from {font_path}")
                return font_name
            except Exception as e:
                print(f"[FONT] Failed to register {font_path}: {e}")
                continue
    print("[FONT] Falling back to Helvetica")
    return 'Helvetica'

PDF_FONT = register_chinese_font()

# JWT Config
SECRET_KEY = "zyzd-secret-key-2024专业发展智诊系统"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24

# FastAPI App
app = FastAPI(title="专业发展智诊系统", version="3.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Excel storage path
EXCEL_DIR = Path(__file__).parent.parent / "excel"
EXCEL_DIR.mkdir(exist_ok=True)

# ============================================================
# Helpers
# ============================================================
def parse_year_from_filename(filename: str) -> Optional[str]:
    """Parse year from filename like '指标、阈值及数据-2020年.xlsx'"""
    # Remove file extension
    name = filename.rsplit('.', 1)[0] if '.' in filename else filename
    
    # Match patterns like: 2020-2021学年, 2020-2021, 2020年, 2020
    patterns = [
        r'(\d{4})\s*[-~]\s*(\d{4})\s*学年?',  # 2020-2021学年 or 2020-2021
        r'(\d{4})\s*学年?',                    # 2020学年 or 2020年 or 2020
    ]
    
    for pattern in patterns:
        match = re.search(pattern, name)
        if match:
            year1 = match.group(1)
            year2 = match.group(2) if len(match.groups()) > 1 and match.group(2) else str(int(year1) + 1)
            return f"{year1}-{year2}学年"
    
    return None

def get_indicator_meta_db() -> Dict[str, Dict]:
    """Get indicator metadata from database"""
    db = SessionLocal()
    try:
        meta = {}
        for ind in db.query(IndicatorMeta).all():
            meta[ind.indicator_id] = {
                "name": ind.name,
                "weight": ind.weight,
                "unit": ind.unit,
                "method": ind.method,
                "thresholds": {
                    "red": ind.red_threshold,
                    "yellow": ind.yellow_threshold,
                    "green": ind.green_threshold
                },
                "higher_is_better": bool(ind.higher_is_better),
                "format": ind.format
            }
        return meta
    finally:
        db.close()

def get_level_value(val: float, ind_id: str, ind_meta: Dict) -> str:
    """Get warning level for an indicator value.
    
    Four levels: red (danger) < yellow (warning) < blue (attention) < green (good)
    Thresholds in database are stored as decimals (0.85 = 85%).
    Values should already be converted to decimal format during import.
    """
    thresholds = ind_meta.get("thresholds", {})
    fmt = ind_meta.get("format", "pct")
    
    # Ensure value is in correct format
    if fmt == "pct" and val > 1:
        val = val / 100.0
    
    if ind_id == "X2":  # 生师比 - lower is better (ratio format)
        # For X2: lower is better, so order is reversed
        # red > yellow > blue > green (values)
        green_thresh = thresholds.get("green", 18)
        blue_thresh = thresholds.get("blue", 20)
        yellow_thresh = thresholds.get("yellow", 22)
        red_thresh = thresholds.get("red", 25)
        
        if val <= green_thresh:
            return "green"
        elif val <= blue_thresh:
            return "blue"
        elif val <= yellow_thresh:
            return "yellow"
        else:
            return "red"
    
    # For indicators where higher is better
    # red < yellow < blue < green (values)
    red_thresh = thresholds.get("red", 0)
    yellow_thresh = thresholds.get("yellow", 0.5)
    blue_thresh = thresholds.get("blue", 0.7)
    green_thresh = thresholds.get("green", 0.9)
    
    # Ensure thresholds are in correct format
    if fmt == "pct":
        if red_thresh > 1:
            red_thresh = red_thresh / 100.0
        if yellow_thresh > 1:
            yellow_thresh = yellow_thresh / 100.0
        if blue_thresh > 1:
            blue_thresh = blue_thresh / 100.0
        if green_thresh > 1:
            green_thresh = green_thresh / 100.0
    
    if val >= green_thresh:
        return "green"
    elif val >= blue_thresh:
        return "blue"
    elif val >= yellow_thresh:
        return "yellow"
    else:
        return "red"

def format_value(val: float, ind_id: str, fmt: str) -> str:
    """Format a value for display."""
    if fmt == "pct":
        return f"{val*100:.1f}%"
    elif fmt == "ratio":
        return f"{val:.1f}"
    elif fmt == "days":
        return f"{val:.0f}天"
    elif fmt == "num":
        return f"{val:.2f}"
    else:
        return f"{val:.2f}"

# ============================================================
# JWT Helpers
# ============================================================
def create_token(data: dict) -> str:
    expire = datetime.utcnow() + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode = data.copy()
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def verify_token(token: str) -> Optional[dict]:
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

# ============================================================
# Auth Models
# ============================================================
class LoginRequest(BaseModel):
    username: str
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    username: str

# ============================================================
# Auth Endpoints
# ============================================================
@app.post("/api/auth/login", response_model=TokenResponse)
async def login(req: LoginRequest):
    if req.username == "admin" and req.password == "admin123":
        token = create_token({"sub": req.username, "role": "admin"})
        return TokenResponse(access_token=token, username=req.username)
    raise HTTPException(status_code=401, detail="用户名或密码错误")

@app.post("/api/auth/logout")
async def logout():
    return {"message": "已退出登录"}

@app.get("/api/auth/me")
async def me(token: str = None):
    if not token:
        raise HTTPException(status_code=401, detail="未登录")
    payload = verify_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="token已过期")
    return {"username": payload.get("sub"), "role": payload.get("role")}

# ============================================================
# Import Endpoint
# ============================================================
@app.post("/api/import")
async def import_excel(
    file: UploadFile = File(...),
    year: str = Form(None)
):
    """Upload and import Excel file"""
    import traceback
    
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="只支持 Excel 文件")
        
        # Parse year from filename if not provided
        if not year:
            year = parse_year_from_filename(file.filename)
        
        if not year:
            raise HTTPException(status_code=400, detail="无法从文件名解析年份，请手动指定")
        
        # Ensure excel directory exists
        EXCEL_DIR.mkdir(parents=True, exist_ok=True)
        
        # Save file
        file_path = EXCEL_DIR / f"{year}_{file.filename}"
        with open(file_path, "wb") as f:
            shutil.copyfileobj(file.file, f)
        
        # Parse Excel
        wb = load_workbook(file_path, data_only=True)
        sheet_names = list(wb.sheetnames)
        
        if len(sheet_names) < 1:
            raise HTTPException(status_code=400, detail="Excel文件至少需要1个工作表")
        
        # Process all sheets as major data (starting from first sheet)
        majors_data = []
        
        for idx in range(0, min(16, len(sheet_names))):
            ws = wb[sheet_names[idx]]
            rows = list(ws.iter_rows(values_only=True))
            
            major_name = sheet_names[idx]
            indicators = {}
            
            # Read rows 1-15 for indicators
            for row_idx in range(1, min(16, len(rows))):
                row = rows[row_idx]
                if row and len(row) >= 5 and row[0] is not None:
                    try:
                        ind_num = int(row[0])
                        if ind_num > 15:
                            continue
                        ind_id = f"X{ind_num}"
                        raw_val = row[3] if row[3] is not None else 0
                        # Handle empty or non-numeric values
                        if raw_val is None or raw_val == '':
                            raw_val = 0
                        
                        val = float(raw_val)
                        
                        # Convert percentage values (>=1) to decimal format (0-1)
                        # For percentage indicators (X1, X3-X13), if value > 1, assume it's 0-100 format and convert to 0-1
                        if ind_id in ['X1', 'X3', 'X4', 'X5', 'X6', 'X7', 'X8', 'X9', 'X10', 'X11', 'X12', 'X13']:
                            if val > 1:
                                val = val / 100.0
                        
                        indicators[ind_id] = val
                    except Exception:
                        pass
            
            majors_data.append({
                "name": major_name,
                "indicators": indicators
            })
        
        if not majors_data:
            raise HTTPException(status_code=400, detail="未找到有效的专业数据")
        
        # Import to database
        success = import_excel_data(year, majors_data)
        
        if success:
            return {"success": True, "message": f"成功导入 {year} 数据，共 {len(majors_data)} 个专业"}
        else:
            raise HTTPException(status_code=500, detail="数据导入失败，请检查数据库连接")
            
    except HTTPException:
        raise
    except Exception as e:
        error_detail = f"导入错误: {str(e)}\n{traceback.format_exc()}"
        print(error_detail)
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")

@app.get("/api/years")
async def get_available_years():
    """Get all available years"""
    years = get_years()
    return {"years": years, "default": years[-1] if years else None}

# ============================================================
# Dashboard Endpoints
# ============================================================
@app.get("/api/dashboard")
async def get_dashboard(year: str = None):
    """Get dashboard overview"""
    years = get_years()
    if not years:
        raise HTTPException(status_code=404, detail="暂无数据，请先导入")
    
    target_year = year or years[-1]
    db_data = get_year_data(target_year)
    
    if not db_data:
        raise HTTPException(status_code=404, detail="年份不存在")
    
    meta = db_data["meta"]
    year_data = db_data["data"].get(target_year, {})
    ind_dict = {ind["id"]: ind for ind in meta["indicators"]}
    
    # Calculate summary
    total_red = total_yellow = total_blue = total_green = 0
    majors_list = []
    
    for m in meta["majors"]:
        mid = m["id"]
        mdata = year_data.get(mid, {})
        
        counts = {"red": 0, "yellow": 0, "blue": 0, "green": 0}
        details = {"red": [], "yellow": [], "blue": [], "green": []}
        
        for ind in meta["indicators"]:
            ind_id = ind["id"]
            val = mdata.get(ind_id, 0)
            level = get_level_value(val, ind_id, ind)
            counts[level] += 1
            details[level].append(ind["name"])
        
        total_red += counts["red"]
        total_yellow += counts["yellow"]
        total_blue += counts["blue"]
        total_green += counts["green"]
        
        health_score = (counts["green"] * 1.0 + counts["blue"] * 0.8 + 
                       counts["yellow"] * 0.5 + counts["red"] * 0) / max(len(meta["indicators"]), 1)
        
        # Calculate composite score based on indicator weights
        total_score = 0
        max_possible_score = 0
        for ind in meta["indicators"]:
            ind_id = ind["id"]
            val = mdata.get(ind_id, 0)
            level = get_level_value(val, ind_id, ind)
            weight = ind.get("weight", 1)
            max_possible_score += weight * 100
            if level == "green":
                total_score += weight * 100
            elif level == "blue":
                total_score += weight * 85
            elif level == "yellow":
                total_score += weight * 60
            else:  # red
                total_score += weight * 30
        
        composite_score = round(total_score / max(max_possible_score, 1) * 100, 1) if max_possible_score > 0 else 0
        
        majors_list.append({
            "id": mid,
            "name": m["name"],
            "fullName": m["fullName"],
            "counts": counts,
            "details": details,
            "healthScore": round(health_score * 100, 1),
            "score": composite_score
        })
    
    majors_list.sort(key=lambda x: x["score"], reverse=True)
    ranking = [{"id": m["id"], "name": m["name"], "healthScore": m["healthScore"], "score": m["score"]} for m in majors_list]
    
    return {
        "year": target_year,
        "years": years,
        "summary": {
            "totalMajors": len(meta["majors"]),
            "red": total_red,
            "yellow": total_yellow,
            "blue": total_blue,
            "green": total_green,
        },
        "majors": majors_list,
        "ranking": ranking
    }

@app.get("/api/major/{major_id}")
async def get_major_detail(major_id: str, year: str = None):
    """Get detailed data for a specific major"""
    years = get_years()
    if not years:
        raise HTTPException(status_code=404, detail="暂无数据")
    
    target_year = year or years[-1]
    db_data = get_year_data(target_year)
    
    if not db_data:
        raise HTTPException(status_code=404, detail="年份不存在")
    
    meta = db_data["meta"]
    year_data = db_data["data"].get(target_year, {})
    mdata = year_data.get(major_id, {})
    
    major_meta = next((m for m in meta["majors"] if m["id"] == major_id), None)
    if not major_meta:
        raise HTTPException(status_code=404, detail="专业不存在")
    
    indicators = []
    ind_dict = {ind["id"]: ind for ind in meta["indicators"]}
    
    for ind in meta["indicators"]:
        ind_id = ind["id"]
        val = mdata.get(ind_id, 0)
        level = get_level_value(val, ind_id, ind_dict)
        
        indicators.append({
            "id": ind_id,
            "name": ind["name"],
            "value": val,
            "level": level,
            "trend": "stable",
            "unit": ind.get("unit", ""),
            "format": ind.get("format", "num"),
            "weight": ind.get("weight", 0)
        })
    
    return {
        "majorId": major_id,
        "majorName": major_meta["name"],
        "year": target_year,
        "indicators": indicators,
        "years": years
    }

@app.get("/api/compare")
async def get_compare(majors: str = None, year: str = None):
    """Get radar chart comparison data"""
    years = get_years()
    if not years:
        raise HTTPException(status_code=404, detail="暂无数据")
    
    target_year = year or years[-1]
    db_data = get_year_data(target_year)
    
    if not db_data:
        raise HTTPException(status_code=404, detail="年份不存在")
    
    meta = db_data["meta"]
    year_data = db_data["data"].get(target_year, {})
    ind_dict = {ind["id"]: ind for ind in meta["indicators"]}
    
    if majors:
        major_ids = majors.split(",")
    else:
        major_ids = [m["id"] for m in meta["majors"]]
    
    core_ids = [f"X{i}" for i in range(1, 16)]
    
    compare_data = []
    for mid in major_ids:
        mdata = year_data.get(mid, {})
        major_meta = next((m for m in meta["majors"] if m["id"] == mid), None)
        name = major_meta["name"] if major_meta else mid
        
        scores = []
        for ind_id in core_ids:
            val = mdata.get(ind_id, 0)
            ind = ind_dict.get(ind_id, {})
            fmt = ind.get("format", "num")
            
            if fmt == "pct":
                score = val * 100
            elif fmt == "ratio":
                score = max(0, min(100, (22 - val) / (22 - 18) * 100))
            elif fmt == "days":
                score = min(val / 30 * 100, 100)
            else:
                score = val * 100
            
            scores.append(round(score, 1))
        
        compare_data.append({"id": mid, "name": name, "scores": scores})
    
    labels = [ind_dict[i]["name"] for i in core_ids if i in ind_dict]
    
    return {
        "year": target_year,
        "indicators": [{"id": i, "name": n} for i, n in zip(core_ids, labels)],
        "majors": compare_data
    }

@app.get("/api/ranking")
async def get_ranking(year: str = None, indicator: str = None):
    """Get ranking data"""
    years = get_years()
    if not years:
        raise HTTPException(status_code=404, detail="暂无数据")
    
    target_year = year or years[-1]
    db_data = get_year_data(target_year)
    
    if not db_data:
        raise HTTPException(status_code=404, detail="年份不存在")
    
    meta = db_data["meta"]
    year_data = db_data["data"].get(target_year, {})
    
    def normalize_value(val, ind_format):
        if ind_format == "pct":
            return val * 100
        elif ind_format == "ratio":
            return val
        elif ind_format == "days":
            return val
        else:
            return val * 100
    
    rankings = []
    for m in meta["majors"]:
        mid = m["id"]
        mdata = year_data.get(mid, {})
        
        if indicator:
            val = mdata.get(indicator, 0)
            ind_meta = next((i for i in meta["indicators"] if i["id"] == indicator), None)
            if ind_meta:
                ind_format = ind_meta.get("format", "pct")
                val = normalize_value(val, ind_format)
        else:
            counts = {"red": 0, "yellow": 0, "blue": 0, "green": 0}
            ind_dict = {i["id"]: i for i in meta["indicators"]}
            for ind_id, ind in ind_dict.items():
                val = mdata.get(ind_id, 0)
                level = get_level_value(val, ind_id, ind)
                counts[level] += 1
            
            # Calculate weighted composite score
            total_score = 0
            max_possible_score = 0
            for ind_id, ind in ind_dict.items():
                val = mdata.get(ind_id, 0)
                level = get_level_value(val, ind_id, ind)
                weight = ind.get("weight", 1)
                max_possible_score += weight * 100
                if level == "green":
                    total_score += weight * 100
                elif level == "blue":
                    total_score += weight * 85
                elif level == "yellow":
                    total_score += weight * 60
                else:  # red
                    total_score += weight * 30
            
            composite_score = round(total_score / max(max_possible_score, 1) * 100, 1) if max_possible_score > 0 else 0
            val = composite_score
        
        rankings.append({"id": mid, "name": m["name"], "value": round(val, 2), "score": round(val, 2)})
    
    higher_is_better = True
    if indicator:
        ind_meta = next((i for i in meta["indicators"] if i["id"] == indicator), None)
        higher_is_better = ind_meta.get("higher_is_better", True) if ind_meta else True
    
    rankings.sort(key=lambda x: x["value"], reverse=higher_is_better)
    for i, r in enumerate(rankings):
        r["rank"] = i + 1
    
    return {"year": target_year, "indicator": indicator, "rankings": rankings}

@app.get("/api/indicator/bar")
async def get_indicator_bar(indicator_id: str = None, year: str = None):
    """Get bar chart data for a specific indicator"""
    years = get_years()
    if not years:
        raise HTTPException(status_code=404, detail="暂无数据")
    
    target_year = year or years[-1]
    db_data = get_year_data(target_year)
    
    if not db_data:
        raise HTTPException(status_code=404, detail="年份不存在")
    
    meta = db_data["meta"]
    year_data = db_data["data"].get(target_year, {})
    
    def normalize_value(val, ind_format):
        if ind_format == "pct":
            return val * 100
        elif ind_format == "ratio":
            return val
        elif ind_format == "days":
            return val
        else:
            return val * 100
    
    if indicator_id:
        ind_meta = next((i for i in meta["indicators"] if i["id"] == indicator_id), None)
        if not ind_meta:
            raise HTTPException(status_code=404, detail="指标不存在")
        
        ind_format = ind_meta.get("format", "pct")
        data = []
        for m in meta["majors"]:
            mid = m["id"]
            mdata = year_data.get(mid, {})
            val = mdata.get(indicator_id, 0)
            normalized_val = normalize_value(val, ind_format)
            data.append({
                "majorId": mid,
                "majorName": m["name"],
                "value": normalized_val,
                "rawValue": val,
                "level": get_level_value(val, indicator_id, {indicator_id: ind_meta}),
                "format": ind_format
            })
        
        reverse_sort = ind_meta.get("higher_is_better", True)
        data.sort(key=lambda x: x["value"], reverse=reverse_sort)
        return {
            "year": target_year,
            "indicator": {"id": indicator_id, "name": ind_meta["name"], "format": ind_format},
            "data": data
        }
    else:
        all_data = {}
        for ind in meta["indicators"]:
            ind_id = ind["id"]
            ind_format = ind.get("format", "pct")
            items = []
            for m in meta["majors"]:
                mid = m["id"]
                mdata = year_data.get(mid, {})
                val = mdata.get(ind_id, 0)
                normalized_val = normalize_value(val, ind_format)
                items.append({
                    "majorId": mid,
                    "majorName": m["name"],
                    "value": normalized_val,
                    "rawValue": val,
                    "level": get_level_value(val, ind_id, {ind_id: ind}),
                    "format": ind_format
                })
            
            # Calculate score for each item based on level
            for item in items:
                level = item["level"]
                if level == "green":
                    item["score"] = 100
                elif level == "blue":
                    item["score"] = 85
                elif level == "yellow":
                    item["score"] = 60
                else:  # red
                    item["score"] = 30
            
            # Sort by score descending
            items.sort(key=lambda x: x["score"], reverse=True)
            all_data[ind_id] = {"name": ind["name"], "format": ind_format, "items": items}
        
        return {"year": target_year, "data": all_data}

# ============================================================
# Trends Endpoint
# ============================================================
@app.get("/api/major/{major_id}/trends")
async def get_major_trends(major_id: str):
    """Get trend data for a major across all years"""
    years = get_years()
    if not years:
        raise HTTPException(status_code=404, detail="暂无数据")
    
    # Sort years in ascending order (oldest first) for trend display
    years = sorted(years)
    
    # Get all year data
    all_years_data = {}
    for year in years:
        db_data = get_year_data(year)
        if db_data and year in db_data["data"]:
            all_years_data[year] = db_data["data"][year]
    
    if not all_years_data:
        raise HTTPException(status_code=404, detail="无趋势数据")
    
    # Get metadata from most recent year
    latest_year = years[-1]
    db_data = get_year_data(latest_year)
    meta = db_data["meta"]
    
    major_meta = next((m for m in meta["majors"] if m["id"] == major_id), None)
    if not major_meta:
        raise HTTPException(status_code=404, detail="专业不存在")
    
    trends = []
    ind_dict = {ind["id"]: ind for ind in meta["indicators"]}
    
    for ind in meta["indicators"]:
        ind_id = ind["id"]
        values = []
        
        for year in years:
            year_data = all_years_data.get(year, {})
            mdata = year_data.get(major_id, {})
            values.append(mdata.get(ind_id, 0))
        
        # Calculate trend slope
        n = len(values)
        if n >= 2:
            x_mean = (n - 1) / 2
            y_mean = sum(values) / n
            num = sum((i - x_mean) * (values[i] - y_mean) for i in range(n))
            den = sum((i - x_mean) ** 2 for i in range(n))
            slope = num / den if den != 0 else 0
        else:
            slope = 0
        
        trends.append({
            "id": ind_id,
            "name": ind["name"],
            "values": values,
            "slope": round(slope, 4),
            "level": get_level_value(values[-1], ind_id, ind_dict) if values else "green",
            "format": ind.get("format", "num"),
            "unit": ind.get("unit", "")
        })
    
    return {"years": years, "trends": trends, "majorName": major_meta["name"]}

# ============================================================
# Warnings Endpoint
# ============================================================
@app.get("/api/warnings")
async def get_warnings(year: str = None):
    """Get all warning items"""
    years = get_years()
    if not years:
        raise HTTPException(status_code=404, detail="暂无数据")
    
    target_year = year or years[-1]
    # Sort years chronologically
    sorted_years = sorted(years)
    
    db_data = get_year_data(target_year)
    
    if not db_data:
        raise HTTPException(status_code=404, detail="年份不存在")
    
    meta = db_data["meta"]
    year_data = db_data["data"].get(target_year, {})
    ind_dict = {ind["id"]: ind for ind in meta["indicators"]}
    
    # Get previous year data
    prev_year_data = {}
    target_year_idx = sorted_years.index(target_year) if target_year in sorted_years else -1
    if target_year_idx > 0:
        prev_year = sorted_years[target_year_idx - 1]
        prev_db_data = get_year_data(prev_year)
        if prev_db_data and prev_year in prev_db_data["data"]:
            prev_year_data = prev_db_data["data"][prev_year]
    
    warnings_list = []
    
    for m in meta["majors"]:
        mid = m["id"]
        mdata = year_data.get(mid, {})
        
        for ind in meta["indicators"]:
            ind_id = ind["id"]
            val = mdata.get(ind_id, 0)
            level = get_level_value(val, ind_id, ind_dict)
            
            if level in ("red", "yellow", "blue"):
                # Calculate change from previous year
                change_val = None
                has_prev_data = prev_year_data and mid in prev_year_data and ind_id in prev_year_data[mid]
                if has_prev_data:
                    prev_val = prev_year_data[mid][ind_id]
                    fmt = ind.get("format", "pct")
                    if fmt == "pct":
                        # For percentage values, show change in percentage points
                        # Both current and previous values are stored as decimals (e.g., 0.85 for 85%)
                        change_val = (val - prev_val) * 100  # Convert to percentage points
                    elif fmt == "ratio":
                        # For ratio values (like student-teacher ratio), show absolute change
                        change_val = val - prev_val
                    elif fmt == "days":
                        # For days values, show absolute change
                        change_val = val - prev_val
                    else:
                        # For other numeric values
                        change_val = val - prev_val
                
                warnings_list.append({
                    "majorId": mid,
                    "majorName": m["name"],
                    "indicatorId": ind_id,
                    "indicatorName": ind["name"],
                    "value": val,
                    "level": level,
                    "change": change_val,
                    "format": ind.get("format", "num"),
                    "unit": ind.get("unit", "")
                })
    
    # Sort: red first, then yellow, then blue
    warnings_list.sort(key=lambda x: (0 if x["level"] == "red" else 1 if x["level"] == "yellow" else 2, x["majorName"]))
    
    return {"year": target_year, "warnings": warnings_list}

# ============================================================
# Report Endpoints
# ============================================================
@app.get("/api/report/{major_id}")
async def generate_report(major_id: str, year: str = None):
    """Generate comprehensive diagnostic report for a major"""
    years = get_years()
    if not years:
        raise HTTPException(status_code=404, detail="暂无数据")
    
    target_year = year or years[-1]
    db_data = get_year_data(target_year)
    
    if not db_data:
        raise HTTPException(status_code=404, detail="年份不存在")
    
    meta = db_data["meta"]
    year_data = db_data["data"].get(target_year, {})
    mdata = year_data.get(major_id, {})
    
    major_meta = next((m for m in meta["majors"] if m["id"] == major_id), None)
    if not major_meta:
        raise HTTPException(status_code=404, detail="专业不存在")
    
    ind_dict = {ind["id"]: ind for ind in meta["indicators"]}
    
    # Categorize indicators
    red_items = []
    yellow_items = []
    blue_items = []
    green_items = []
    
    # Calculate average for each indicator across all majors
    indicator_averages = {}
    for ind in meta["indicators"]:
        ind_id = ind["id"]
        values = []
        for m in meta["majors"]:
            mid = m["id"]
            mdata_all = year_data.get(mid, {})
            val = mdata_all.get(ind_id, 0)
            if val is not None:
                values.append(val)
        if values:
            avg_val = sum(values) / len(values)
            # Convert to score based on level
            level_avg = get_level_value(avg_val, ind_id, ind_dict)
            if level_avg == "green":
                score_avg = 100
            elif level_avg == "blue":
                score_avg = 85
            elif level_avg == "yellow":
                score_avg = 60
            else:
                score_avg = 30
            indicator_averages[ind_id] = {
                "value": avg_val,
                "score": score_avg,
                "level": level_avg
            }
    
    for ind in meta["indicators"]:
        ind_id = ind["id"]
        val = mdata.get(ind_id, 0)
        level = get_level_value(val, ind_id, ind_dict)
        
        # Calculate score based on level
        score = 0
        if level == "green":
            score = 100
        elif level == "blue":
            score = 85
        elif level == "yellow":
            score = 60
        else:  # red
            score = 30
        
        item = {
            "id": ind_id,
            "name": ind["name"],
            "value": val,
            "score": score,
            "level": level,
            "trend": "stable",
            "change": None,
            "unit": ind.get("unit", ""),
            "format": ind.get("format", "num")
        }
        
        if level == "red":
            red_items.append(item)
        elif level == "yellow":
            yellow_items.append(item)
        elif level == "blue":
            blue_items.append(item)
        else:
            green_items.append(item)
    
    # Calculate health score
    total = len(meta["indicators"])
    health_score = (len(green_items) * 100 + len(blue_items) * 80 + 
                   len(yellow_items) * 50 + len(red_items) * 0) / max(total, 1)
    
    # Generate report text - NEW FORMAT
    report_lines = []
    report_lines.append(f"{'='*50}")
    report_lines.append(f"【{major_meta['name']}】专业发展智诊报告")
    report_lines.append(f"生成时间：{datetime.now().strftime('%Y年%m月%d日 %H:%M')}")
    report_lines.append(f"数据年度：{target_year}")
    report_lines.append(f"{'='*50}")
    report_lines.append("")
    
    # 一、总体评价 - 新增排名和健康度变化
    report_lines.append("一、总体评价")
    report_lines.append(f"本专业共监测{len(meta['indicators'])}项核心指标，")
    report_lines.append(f"其中绿色指标{len(green_items)}项、蓝色关注指标{len(blue_items)}项、")
    report_lines.append(f"黄色预警指标{len(yellow_items)}项、红色预警指标{len(red_items)}项。")
    report_lines.append(f"综合健康度得分：{health_score:.1f}分。")
    
    # 新增：专业排名和评价
    # 计算专业排名
    all_majors_scores = []
    for m in meta["majors"]:
        mid = m["id"]
        mdata_check = year_data.get(mid, {})
        m_green = m_blue = m_yellow = m_red = 0
        for ind_check in meta["indicators"]:
            ind_id_check = ind_check["id"]
            val_check = mdata_check.get(ind_id_check, 0)
            level_check = get_level_value(val_check, ind_id_check, ind_check)
            if level_check == "green":
                m_green += 1
            elif level_check == "blue":
                m_blue += 1
            elif level_check == "yellow":
                m_yellow += 1
            else:
                m_red += 1
        m_health = (m_green * 100 + m_blue * 80 + m_yellow * 50 + m_red * 0) / max(len(meta["indicators"]), 1)
        all_majors_scores.append({"id": mid, "name": m["name"], "score": m_health})
    
    all_majors_scores.sort(key=lambda x: x["score"], reverse=True)
    rank = next((i + 1 for i, m in enumerate(all_majors_scores) if m["id"] == major_id), len(all_majors_scores))
    total_majors = len(all_majors_scores)
    
    # 确定评价等级
    if health_score >= 80:
        evaluation = "优秀"
        development_status = "良好"
        stability = "稳定"
    elif health_score >= 60:
        evaluation = "良好"
        development_status = "正常"
        stability = "稳定"
    elif health_score >= 40:
        evaluation = "正常"
        development_status = "较差"
        stability = "不稳定"
    else:
        evaluation = "不合格"
        development_status = "较差"
        stability = "不稳定"
    
    # 计算与上学期的对比（如果有历史数据）
    prev_year = None
    prev_health_score = None
    prev_rank = None
    years_list = get_years()
    if target_year in years_list:
        year_idx = years_list.index(target_year)
        if year_idx > 0:
            prev_year = years_list[year_idx - 1]
            # 获取上年数据计算健康度
            prev_db_data = get_year_data(prev_year)
            if prev_db_data:
                prev_meta = prev_db_data["meta"]
                prev_year_data = prev_db_data["data"].get(prev_year, {})
                prev_mdata = prev_year_data.get(major_id, {})
                
                prev_green = prev_blue = prev_yellow = prev_red = 0
                for ind in prev_meta["indicators"]:
                    ind_id = ind["id"]
                    val = prev_mdata.get(ind_id, 0)
                    level = get_level_value(val, ind_id, ind)
                    if level == "green":
                        prev_green += 1
                    elif level == "blue":
                        prev_blue += 1
                    elif level == "yellow":
                        prev_yellow += 1
                    else:
                        prev_red += 1
                
                prev_health_score = (prev_green * 100 + prev_blue * 80 + prev_yellow * 50 + prev_red * 0) / max(len(prev_meta["indicators"]), 1)
                
                # 计算上年排名
                prev_all_majors = []
                for m in prev_meta["majors"]:
                    mid = m["id"]
                    mdata_p = prev_year_data.get(mid, {})
                    p_green = p_blue = p_yellow = p_red = 0
                    for ind_p in prev_meta["indicators"]:
                        ind_id_p = ind_p["id"]
                        val_p = mdata_p.get(ind_id_p, 0)
                        level_p = get_level_value(val_p, ind_id_p, ind_p)
                        if level_p == "green": p_green += 1
                        elif level_p == "blue": p_blue += 1
                        elif level_p == "yellow": p_yellow += 1
                        else: p_red += 1
                    p_health = (p_green * 100 + p_blue * 80 + p_yellow * 50 + p_red * 0) / max(len(prev_meta["indicators"]), 1)
                    prev_all_majors.append({"id": mid, "score": p_health})
                prev_all_majors.sort(key=lambda x: x["score"], reverse=True)
                prev_rank = next((i + 1 for i, m in enumerate(prev_all_majors) if m["id"] == major_id), len(prev_all_majors))
    
    score_change_text = ""
    rank_change_text = ""
    if prev_health_score is not None:
        score_diff = health_score - prev_health_score
        if score_diff > 0:
            score_change_text = f"增加{score_diff:.1f}分"
        elif score_diff < 0:
            score_change_text = f"减少{abs(score_diff):.1f}分"
        else:
            score_change_text = "持平"
        
        if prev_rank is not None:
            rank_diff = prev_rank - rank
            if rank_diff > 0:
                rank_change_text = f"上升{rank_diff}个名次"
            elif rank_diff < 0:
                rank_change_text = f"下降{abs(rank_diff)}个名次"
            else:
                rank_change_text = "排名不变"
    
    # 输出新增段落
    report_lines.append(f"专业综合排名在第{rank}名（共{total_majors}个专业）。")
    report_lines.append(f"总体评价：{evaluation}。该专业整体发展状况{development_status}，各项指标表现{stability}。")
    
    if score_change_text and rank_change_text:
        report_lines.append(f"本专业较上一学年综合健康度得分{score_change_text}，排名{rank_change_text}。")
    
    report_lines.append("")
    
    # 二、各指标表现情况（原：各指标预警情况）
    report_lines.append("二、各指标表现情况")
    
    # 计算最高分、最低分、未得分指标数量
    max_count = 0
    min_count = 0
    zero_count = 0
    
    # 获取所有专业的数据来计算
    all_majors_data = {}
    for m in meta["majors"]:
        mid = m["id"]
        mdata_all = year_data.get(mid, {})
        all_majors_data[mid] = mdata_all
    
    # 对每个指标，检查当前专业是否是最高/最低
    for ind in meta["indicators"]:
        ind_id = ind["id"]
        current_val = mdata.get(ind_id, 0)
        
        # 获取所有专业该指标的值
        all_values = []
        for mid, mdata_all in all_majors_data.items():
            val = mdata_all.get(ind_id, 0)
            if val is not None and val > 0:
                all_values.append(val)
        
        if len(all_values) > 0:
            max_val = max(all_values)
            min_val = min(all_values)
            
            # 检查是否是最高分
            if current_val > 0 and abs(current_val - max_val) < 0.0001:
                max_count += 1
            # 检查是否是最低分
            if current_val > 0 and abs(current_val - min_val) < 0.0001:
                min_count += 1
        
        # 检查是否未得分（值为0或null）
        if current_val is None or current_val == 0:
            zero_count += 1
    
    report_lines.append(f"{major_meta['name']}共有{max_count}个指标在各专业中取得了最高分，共有{min_count}个指标在各专业中取得了最低分，共有{zero_count}个指标未得分。")
    report_lines.append("")
    
    # 红色预警指标
    red_count = len(red_items)
    report_lines.append(f"红色预警指标（{red_count}项）：")
    if red_items:
        for item in red_items:
            val_str = format_value(item["value"], item["id"], item["format"])
            report_lines.append(f"• {item['name']}：{val_str}，数据不在正常范围内，建议立刻采取行动，扭转表现不佳的态势。")
    report_lines.append(f"本专业在{', '.join([item['name'] for item in red_items[:3]]) if red_items else '暂无'}方面处于劣势，需增强危机意识，立刻分析原因，提出改善举措，主动创新，寻求突破，弥补短板，借鉴表现优异的专业建设经验。")
    report_lines.append("")
    
    # 黄色预警指标
    yellow_count = len(yellow_items)
    report_lines.append(f"黄色预警指标（{yellow_count}项）：")
    if yellow_items:
        # 简化处理：使用在校生满意度作为示例
        student_satisfaction = next((item for item in yellow_items if '满意度' in item['name']), yellow_items[0] if yellow_items else None)
        if student_satisfaction:
            val_str = format_value(student_satisfaction["value"], student_satisfaction["id"], student_satisfaction["format"])
            report_lines.append(f"• 在校生满意度：{val_str}，较上一学年增加/减少了X， 如未加以关注，数据将下滑至异常范围，建议密切关注。")
    report_lines.append("本专业在学生满意度方面表现不佳，需树立全局意识，统筹发展；认清自身不足，深化改革，取长补短，改善现状。")
    report_lines.append("")
    
    # 蓝色关注指标
    blue_count = len(blue_items)
    report_lines.append(f"蓝色关注指标（{blue_count}项）：")
    if blue_items:
        for item in blue_items[:2]:  # 最多显示2个
            val_str = format_value(item["value"], item["id"], item["format"])
            report_lines.append(f"• {item['name']}：{val_str}，正常但有负向波动，需分析波动原因，避免持续走低。")
    report_lines.append("本专业在社会吸引力、学生考取技能证书和学生专业认可度方面表现优异，但较上一学年呈现下降趋势，需及时分析下降原因，保持平稳发展趋势。")
    report_lines.append("")
    
    # 绿色健康指标
    green_count = len(green_items)
    report_lines.append(f"绿色健康指标（{green_count}项）：")
    if green_items:
        for item in green_items[:2]:  # 最多显示2个
            val_str = format_value(item["value"], item["id"], item["format"])
            report_lines.append(f"• {item['name']}：{val_str}，趋势健康。")
    report_lines.append("本专业在生师配比、课程教学效果、学生毕业就业和产教融合方面表现优异，需继续保持。")
    report_lines.append("")
    
    # 三、综合改进建议
    report_lines.append("三、综合改进建议")
    
    if red_items:
        red_names = ', '.join([item['name'] for item in red_items[:3]])
        report_lines.append(f"1. 寻求突破：提升红色异常指标{red_names}，分析原因，精准突破；借鉴其他专业宝贵经验，明确特色化发展路径。")
    
    if yellow_items:
        yellow_names = ', '.join([item['name'] for item in yellow_items[:3]])
        report_lines.append(f"2.重点提升：改善黄色预警指标{yellow_names}，制定针对性改进计划，避免现状恶化。")
    
    if blue_items:
        blue_names = ', '.join([item['name'] for item in blue_items[:3]])
        report_lines.append(f"3. 持续关注：保持蓝色指标{blue_names}的稳定性，防止进一步下滑。")
    
    if green_items:
        green_names = ', '.join([item['name'] for item in green_items[:3]])
        report_lines.append(f"4.稳步发展：增强绿色指标{green_names}的稳定性，及时总结建设经验并迁移至其他指标，实现本专业全面统筹发展。")
    
    report_lines.append("")
    report_lines.append(f"{'='*50}")
    report_lines.append("报告由专业发展智诊系统自动生成")
    report_lines.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    report_text = "\n".join(report_lines)
    
    # Build sorted items list (red -> yellow -> blue -> green)
    sorted_items = []
    for item in red_items + yellow_items + blue_items + green_items:
        sorted_items.append(item)
    
    return {
        "majorId": major_id,
        "majorName": major_meta["name"],
        "year": target_year,
        "healthScore": round(health_score, 1),
        "red": red_items,
        "yellow": yellow_items,
        "blue": blue_items,
        "green": green_items,
        "allItems": sorted_items,
        "indicatorAverages": indicator_averages,
        "reportText": report_text
    }

@app.get("/api/report/{major_id}/pdf")
async def download_report_pdf(major_id: str, year: str = None, token: str = None):
    """Download report as PDF"""
    # Verify token
    if not token:
        raise HTTPException(status_code=401, detail="Missing token")
    payload = verify_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    
    years = get_years()
    if not years:
        raise HTTPException(status_code=404, detail="暂无数据")
    
    target_year = year or years[-1]
    db_data = get_year_data(target_year)
    
    if not db_data:
        raise HTTPException(status_code=404, detail="年份不存在")
    
    meta = db_data["meta"]
    year_data = db_data["data"].get(target_year, {})
    mdata = year_data.get(major_id, {})
    
    major_meta = next((m for m in meta["majors"] if m["id"] == major_id), None)
    if not major_meta:
        raise HTTPException(status_code=404, detail="专业不存在")
    
    # Get report data
    report_data = await generate_report(major_id, target_year)
    report_text = report_data["reportText"]
    
    # Create PDF
    output_dir = Path(__file__).parent.parent / "pdf_reports"
    output_dir.mkdir(exist_ok=True)
    
    pdf_path = output_dir / f"{major_meta['name']}_{target_year}_诊断报告.pdf"
    
    # Build PDF content
    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    # Create Chinese paragraph style
    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontName=PDF_FONT,
        fontSize=11,
        leading=18,
        spaceBefore=6,
        spaceAfter=6
    )
    
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        fontName=PDF_FONT,
        fontSize=16,
        leading=22,
        spaceBefore=12,
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'Heading',
        parent=styles['Heading2'],
        fontName=PDF_FONT,
        fontSize=13,
        leading=18,
        spaceBefore=12,
        spaceAfter=8
    )
    
    story = []
    
    # Add title
    story.append(Paragraph(f"{major_meta['name']} 专业发展诊断报告", title_style))
    story.append(Paragraph(f"数据年度：{target_year}", normal_style))
    story.append(Spacer(1, 0.5*cm))
    
    # Add report text (parse line by line)
    for line in report_text.split('\n'):
        line = line.strip()
        if not line:
            story.append(Spacer(1, 0.3*cm))
            continue
        
        # Check if it's a heading
        if line.startswith('一、') or line.startswith('二、') or line.startswith('三、'):
            story.append(Paragraph(line.replace('一、', '').replace('二、', '').replace('三、', ''), heading_style))
        else:
            story.append(Paragraph(line, normal_style))
    
    # Build PDF
    doc.build(story)
    
    return FileResponse(str(pdf_path), media_type='application/pdf', 
                        filename=f"{major_meta['name']}_{target_year}_诊断报告.pdf")

# ============================================================
# Static Files
# ============================================================
@app.get("/")
async def root():
    html_file = Path(__file__).parent.parent / "frontend" / "index.html"
    return FileResponse(str(html_file))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8089)
